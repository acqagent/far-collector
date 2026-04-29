"""Excel export for the two FAR spreadsheets."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rich.console import Console

import db

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

console = Console()


def style_header(ws, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def autosize(ws, max_widths: dict[int, int]) -> None:
    for c in range(1, ws.max_column + 1):
        max_len = 10
        for row in ws.iter_rows(min_col=c, max_col=c, min_row=1, max_row=ws.max_row, values_only=True):
            v = row[0]
            if v is None:
                continue
            ln = min(len(str(v)), max_widths.get(c, 60))
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[get_column_letter(c)].width = max_len + 2


def _read_only_con():
    """Snapshot the live DuckDB to a temp file and read from there.

    DuckDB enforces single-process file locks even for read-only opens, so when
    the long-running deviation crawler is active we copy the DB and query the
    copy instead of waiting for the writer.
    """
    import shutil
    import tempfile

    import duckdb

    snap = Path(tempfile.gettempdir()) / "collector_export_snapshot.duckdb"
    shutil.copyfile(db.DB, snap)
    return duckdb.connect(str(snap), read_only=True)


def export_provisions() -> Path:
    con = _read_only_con()
    rows = con.execute("""
        SELECT number, title, kind, effective_date, full_text, source_url, scraped_at
        FROM far_provisions_clauses
        ORDER BY number
    """).fetchall()
    con.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "Provisions & Clauses"
    headers = ["Number", "Title", "Type", "Effective Date", "Full Text", "Source URL", "Scraped At"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws, len(headers))
    autosize(ws, {1: 16, 2: 60, 3: 12, 4: 18, 5: 80, 6: 60, 7: 22})
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = WRAP
    out = OUTPUT_DIR / "far_provisions_clauses.xlsx"
    wb.save(out)
    console.print(f"[green]Wrote[/] {out} ({len(rows)} rows)")
    return out


def export_deviations() -> Path:
    con = _read_only_con()
    rows = con.execute("""
        SELECT agency, deviation_number, title,
               effective_date, effective_date_iso, effective_date_kind,
               scope, link, scraped_at
        FROM far_class_deviations
        WHERE agency NOT IN ('DoD','DOD','Department of Defense')
        ORDER BY effective_date_iso DESC NULLS LAST, agency, deviation_number
    """).fetchall()
    con.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "Class Deviations"
    headers = [
        "Agency", "Deviation #", "Title",
        "Effective Date (raw)", "Effective Date (ISO)", "Date Kind",
        "Scope", "Link", "Scraped At",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws, len(headers))
    autosize(ws, {1: 14, 2: 22, 3: 60, 4: 22, 5: 14, 6: 12, 7: 80, 8: 60, 9: 22})
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell.alignment = WRAP
    out = OUTPUT_DIR / "far_class_deviations.xlsx"
    wb.save(out)
    console.print(f"[green]Wrote[/] {out} ({len(rows)} rows)")
    return out


def main(target: str = "all") -> int:
    if target in ("provisions", "all"):
        export_provisions()
    if target in ("deviations", "all"):
        export_deviations()
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else "all"))
